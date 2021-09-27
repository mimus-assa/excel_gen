# import the flask class
from flask import Flask, session, render_template, request, make_response, redirect, flash, url_for, send_from_directory
from flaskext.mysql import MySQL
from flask_bootstrap import Bootstrap
import mysql.connector
# instatiating flask class 
import io
import csv
from openpyxl import load_workbook
import os
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from datetime import datetime
import pandas as pd

app=Flask(__name__)
mysql = MySQL()
 
# configuring MySQL for the web application
app.config['MYSQL_DATABASE_USER'] = 'XXXXXXX'    # default user of MySQL to be replaced with appropriate username
app.config['MYSQL_DATABASE_PASSWORD'] = 'XXXXXX' # default passwrod of MySQL to be replaced with appropriate password
app.config['MYSQL_DATABASE_DB'] = 'XXXXX'  # Database name to be replaced with appropriate database name
app.config['MYSQL_DATABASE_HOST'] = 'localhost' # default database host of MySQL to be replaced with appropriate database host
#initialise mySQL
mysql.init_app(app)
#create connection to access data
conn = mysql.connect()

bootstrap = Bootstrap(app)

@app.route("/error_page")    
def error_page():
    return render_template("404.html")  

@app.route("/")    
def index():
    return render_template("index.html")

@app.route("/download/base", methods=["GET", "POST"])
def Downloads_base():
    mycursor = conn.cursor()
    sql = "SELECT * FROM Servi"
    mycursor.execute(sql)
    result = mycursor.fetchall()
    Servi_DB = pd.DataFrame(list(result))

    now = datetime.now()
    dt_string = now.strftime("%d-%m-%Y-%H-%M-%S")
    NAME = "generated/"+dt_string+"_"+"DB"+".csv"
    uploads = os.path.join(app.root_path, "generated")
    
    Servi_DB.to_csv(NAME, encoding='utf_8_sig')
    return send_from_directory(directory=uploads, filename=dt_string+"_"+"DB"+".csv")

@app.route("/edicion_form/<order>", methods=["GET", "POST"])    
def edicion_form(order):
    try:
        mycursor = conn.cursor()
        sql = "SELECT * FROM Servi WHERE Orden='%s'"%order
        mycursor.execute(sql)
        result = mycursor.fetchall()
    

        if request.method == 'POST':
            Orden = request.form['Orden']
            Dependencia = request.form['Dependencia']
            Atencion = request.form['Atencion']
            Recibio_del_cer = request.form['Recibio_del_cer']
            Entrego_al_cer = request.form['Entrego_al_cer']
            Folio = request.form['Folio']
            Entrada = request.form['Entrada']
            Equipo = request.form['Equipo']
            Marca = request.form['Marca']
            Modelo = request.form['Modelo']
            Serie = request.form['Serie']
            Unidad = request.form['Unidad']
            Delegacion = request.form['Delegacion']
            Accesorio = request.form['Accesorio']
            Falla = request.form['Falla']
            Ticket = request.form['Ticket']
            Status = request.form['Status']
            Reparacion = request.form['Reparacion']
            Salida = request.form['Salida']
            Reparo = request.form['Reparo']
            Servicio = request.form['Servicio']
            Reporte = request.form['Reporte']
            sql2  = "UPDATE Servi SET Orden='%s', Dependencia='%s', Atencion='%s', Recibio_del_cer='%s', Entrego_al_cer='%s', Folio='%s', Entrada='%s', Equipo='%s', Marca='%s', Modelo='%s', Serie='%s', Unidad='%s', Delegacion='%s', Accesorio='%s', Falla='%s', Ticket='%s', Status='%s', Reparacion='%s', Salida='%s', Reparo='%s', Servicio='%s', Reporte='%s' WHERE Orden='%s'" %(Orden, Dependencia, Atencion, Recibio_del_cer, Entrego_al_cer, Folio, Entrada, Equipo, Marca, Modelo, Serie, Unidad, Delegacion, Accesorio, Falla, Ticket, Status, Reparacion, Salida, Reparo, Servicio, Reporte, Orden)
        
            mycursor = conn.cursor()
            mycursor.execute(sql2)
            conn.commit()
            return redirect("/edicion")
    except:
        return redirect("404.html")
    try:
        return render_template("edition_form.html", data=result[0])
    except:
        return redirect("404.html")
@app.route("/edicion", methods=["GET", "POST"])    
def edicion_busqueda():
    try:
        if request.method == 'POST':
            Orden = request.form['Orden']
        
            url_download = "/edicion_form/"+Orden
            return redirect(url_download)
    except:
        return redirect("404.html")
    return render_template("edicion.html")
@app.route("/captura", methods=["GET", "POST"])
def show_signup_form():
    mycursor = conn.cursor()
    sql2 = "SELECT Orden FROM Servi ORDER BY ID DESC LIMIT 1"
    mycursor.execute(sql2)
    result = mycursor.fetchall()
    ordin = result[0]
    numeric_filter = filter(str.isdigit, ordin[0])
    numeric_string = "".join(numeric_filter)
    
    #print("A"+str(int(numeric_string)+1))
    ordin = "A"+str(int(numeric_string)+1)
    #print(ordin)
    if request.method == 'POST':
        Orden = request.form['Orden']
        Dependencia = request.form['Dependencia']
        Atencion = request.form['Atencion']
        Recibio_del_cer = request.form['Recibio_del_cer']
        Entrego_al_cer = request.form['Entrego_al_cer']
        Folio = request.form['Folio']
        Entrada = request.form['Entrada']
        Equipo = request.form['Equipo']
        Marca = request.form['Marca']
        Modelo = request.form['Modelo']
        Serie = request.form['Serie']
        Unidad = request.form['Unidad']
        Delegacion = request.form['Delegacion']
        Accesorio = request.form['Accesorio']
        Falla = request.form['Falla']
        Ticket = request.form['Ticket']
        Status = request.form['Status']
        Reparacion = request.form['Reparacion']
        Salida = request.form['Salida']
        Reparo = request.form['Reparo']
        Servicio = request.form['Servicio']
        Reporte = request.form['Reporte']
        sql = "INSERT INTO Servi ( Orden, Dependencia, Atencion, Recibio_del_cer, Entrego_al_cer, Folio, Entrada, Equipo, Marca, Modelo, Serie, Unidad, Delegacion, Accesorio, Falla, Ticket, Status, Reparacion, Salida, Reparo, Servicio, Reporte) VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        value = (Orden, Dependencia, Atencion, Recibio_del_cer, Entrego_al_cer, Folio, Entrada, Equipo, Marca, Modelo, Serie, Unidad, Delegacion, Accesorio, Falla, Ticket, Status, Reparacion, Salida, Reparo, Servicio, Reporte)
        numeric_filter_ord = filter(str.isdigit, Orden)
        numeric_string_ord = "".join(numeric_filter_ord)
        if int(numeric_string_ord) < int(numeric_string)+1:
            return redirect(url_for('error_page'))
        print(numeric_string_ord, int(numeric_string)+1)
        mycursor.execute(sql, value)
        conn.commit()
        next = request.args.get('next', None)
        
        if next:
            print("here: ",next)
            return redirect(next)
        return redirect(url_for('show_signup_form'))
    return render_template("input.html", data=ordin)

   
    
@app.route("/orden/<order>")    
def xlsx_row(order):
    try:
        mycursor = conn.cursor()
        wb_orden = load_workbook(filename = 'orden.xlsx')
        ws_orden = wb_orden["ORDEN"]
        querry = "SELECT * FROM Servi WHERE Orden='%s'" %order
        mycursor.execute(querry)
        result = mycursor.fetchall()
        ordin = result[0]
        ws_orden["H6"]  = ordin[1]
        ws_orden["B6"],  ws_orden["H7"], ws_orden["H44"],  ws_orden["B10"], ws_orden["B11"],  ws_orden["B12"], ws_orden["B13"], ws_orden["B16"], ws_orden["B17"], ws_orden["B44"], ws_orden["H11"], ws_orden["A19"], ws_orden["B7"], ws_orden["B14"], ws_orden["B15"], ws_orden["H8"] = ordin[2], ordin[6], ordin[7], ordin[8], ordin[9], ordin[10], ordin[11], ordin[12], ordin[13], ordin[20], ordin[21], ordin[22], ordin[3], ordin[14], ordin[15], ordin[16]
        now = datetime.now()
        dt_string = now.strftime("%d-%m-%Y-%H-%M-%S")
        NAME = "generated/"+dt_string+"_"+order+".xlsx"
        wb_orden.save(filename = NAME)
        uploads = os.path.join(app.root_path, "generated")
        return send_from_directory(directory=uploads, filename=dt_string+"_"+order+".xlsx")
    except:
        return render_template("404.html")
    return render_template("404.html")
@app.route("/resguardo/<order>")
def xlsx_res(order):
    try:
        mycursor = conn.cursor()
        wb_orden = load_workbook(filename = 'resg.xlsx')
        ws_orden = wb_orden["Resguardo Individual"]
        querry = "SELECT * FROM Servi WHERE Orden='%s'" %order
        mycursor.execute(querry)
        result = mycursor.fetchall()
        ordin = result[0]
        ws_orden["G6"] = ordin[1]
        ws_orden["B7"],  ws_orden["G7"], ws_orden["G8"],  ws_orden["B11"], ws_orden["B12"],  ws_orden["B13"], ws_orden["B14"], ws_orden["G11"],  ws_orden["B8"], ws_orden["B15"], ws_orden["B16"], ws_orden["B20"], ws_orden["F20"] = ordin[2], ordin[6], ordin[7], ordin[8], ordin[9], ordin[10], ordin[11], ordin[13], ordin[3], ordin[14], ordin[15], ordin[5], ordin[4]
        now = datetime.now()
        dt_string = now.strftime("%d-%m-%Y-%H-%M-%S")
        NAME = "generated/"+"res_"+dt_string+"_"+order+".xlsx"
        wb_orden.save(filename = NAME)
        uploads = os.path.join(app.root_path, "generated")
        return send_from_directory(directory=uploads, filename="res_"+dt_string+"_"+order+".xlsx")
    except:
        return render_template("404.html")
@app.route("/download/orden", methods=["GET", "POST"])
def Downloads_order():
    if request.method == 'POST':
        Orden = request.form['Orden']
        next = request.args.get('next', None)
        if next:
            return redirect(next)
        url_download = "/orden/"+Orden
        print(url_download)
        return redirect(url_download)
    return render_template("downloads.html")
    
@app.route("/download/resguardo", methods=["GET", "POST"])
def Downloads_resg():
    if request.method == 'POST':
        Orden = request.form['Orden']
        next = request.args.get('next', None)
        if next:
            return redirect(next)
        url_download = "/resguardo/"+Orden
        print(url_download)
        return redirect(url_download)
    return render_template("downloads.html")
    
 
    
        
if __name__ == '__main__':
    app.run(host='0.0.0.0', port='5000', debug=True)
